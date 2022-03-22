from openpyxl import load_workbook
from fitter import Fitter

arrivals = load_workbook('arrivals.xlsx')
call_durations = load_workbook('calldurations.xlsx')
speeds = load_workbook('speeds.xlsx')

arrival_time = []
arrival_station = []
row = 2
while value := arrivals.active[f'B{row}'].value:
    arrival_time.append(value)
    arrival_station.append(arrivals.active[f'C{row}'].value)
    assert isinstance(value, (float, int))
    assert isinstance(arrival_station[-1], int)
    row += 1

arrival_interval = [t2 - t1 for t1, t2 in zip(arrival_time[:-1], arrival_time[1:])]

duration = []
row = 2
while value := call_durations.active[f'A{row}'].value:
    duration.append(value)
    assert isinstance(value, (float, int))
    row += 1

speed = []
row = 2
while value := speeds.active[f'A{row}'].value:
    speed.append(value)
    assert isinstance(value, (float, int))
    row += 1

fitter = Fitter(arrival_interval, timeout=10)
fitter.fit()
arrival_interval_dist = fitter.get_best()

fitter = Fitter(arrival_station, timeout=10)
fitter.fit()
arrival_station_dist = fitter.get_best()

fitter = Fitter(duration, timeout=10)
fitter.fit()
duration_dist = fitter.get_best()

fitter = Fitter(speed, timeout=10)
fitter.fit()
speed_dist = fitter.get_best()

print(f'Arrival interval distribution: {arrival_interval_dist}')
print(f'Arrival base station distribution: {arrival_station_dist}')
print(f'Call duration distribution: {duration_dist}')
print(f'Speed distribution: {speed_dist}')

# Output:
# Arrival interval distribution: {'expon': {'loc': 0.0, 'scale': 1.3499896868554284}}
# Arrival base station distribution: {'uniform': {'loc': 0.0, 'scale': 19.0}}
# Call duration distribution: {'expon': {'loc': 0.00706693, 'scale': 119.094604758143}}
# Speed distribution: {'norm': {'loc': 89.98472287, 'scale': 8.216482873655245}}
#
# From here we know that arriving calls uniformly distributed across all base
# stations.
# The data does not provide further information on how it distributed within
# each base station's cell, for now I assume it is also uniform in further
# simulation programs.